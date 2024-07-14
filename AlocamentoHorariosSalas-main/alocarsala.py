import random
from openpyxl import load_workbook

def alocar_salas(planilha_path):
    print(f"Carregando planilha: {planilha_path}")
    
    try:
        # Carrega o workbook existente
        workbook = load_workbook(planilha_path)
    except Exception as e:
        print(f"Erro ao carregar a planilha: {e}")
        return
    
    # Definindo salas disponíveis
    salas = [f"220{i}" for i in range(1, 10)] + [f"230{i}" for i in range(1, 10)] + [f"240{i}" for i in range(1, 10)]
    
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        print(f"Alocando salas na planilha: {sheet_name}")
        
        # Armazena combinações usadas para evitar repetições de salas no mesmo dia e horário
        combinacoes_usadas = set()
        
        # Adiciona uma nova coluna "Sala" se não existir
        if "Sala" not in [cell.value for cell in sheet[1]]:
            sheet.cell(row=1, column=sheet.max_column + 1, value="Sala")
        
        sala_col_idx = sheet.max_column
        
        # Itera sobre as linhas, começando da segunda linha (primeira linha são os cabeçalhos)
        for row in sheet.iter_rows(min_row=2, max_col=sheet.max_column):
            try:
                dia1, hora1, dia2, hora2 = row[0].value, row[2].value, row[1].value, row[3].value
                
                # Pular linhas onde Dia1 e Dia2 estão vazios
                if (not dia1 or not hora1) and (not dia2 or not hora2):
                    print(f"Linha {row[0].row} na planilha {sheet_name} não possui dados de dias ou horas. Ignorando...")
                    continue

                # Limite de tentativas para alocar salas
                max_attempts = 100
                attempts = 0

                # Aloca uma sala para o primeiro dia e hora
                sala1 = None
                if dia1 and hora1:
                    while attempts < max_attempts:
                        sala1 = random.choice(salas)
                        combinacao1 = (dia1, hora1, sala1)
                        if combinacao1 not in combinacoes_usadas:
                            combinacoes_usadas.add(combinacao1)
                            break
                        attempts += 1
                    else:
                        print(f"Não foi possível alocar uma sala para a combinação {dia1}, {hora1} na linha {row[0].row} na planilha {sheet_name}")
                        continue

                attempts = 0
                # Aloca uma sala para o segundo dia e hora, se existir
                sala2 = None
                if dia2 and hora2:
                    while attempts < max_attempts:
                        sala2 = random.choice(salas)
                        combinacao2 = (dia2, hora2, sala2)
                        if combinacao2 not in combinacoes_usadas and combinacao2 != combinacao1:
                            combinacoes_usadas.add(combinacao2)
                            break
                        attempts += 1
                    else:
                        print(f"Não foi possível alocar uma sala para a combinação {dia2}, {hora2} na linha {row[0].row} na planilha {sheet_name}")
                        continue
                
                # Preenche a célula da nova coluna com a sala alocada
                sala_value = ""
                if sala1:
                    sala_value = sala1
                if sala2:
                    if sala_value:
                        sala_value += f", {sala2}"
                    else:
                        sala_value = sala2

                sheet.cell(row=row[0].row, column=sala_col_idx, value=sala_value)
                print(f"Alocando salas para linha {row[0].row} na planilha {sheet_name}: {sala_value}")
            except Exception as e:
                print(f"Erro ao processar a linha {row[0].row} na planilha {sheet_name}: {e}")
    
    try:
        # Salva o workbook com as novas informações
        workbook.save(planilha_path)
        print("Salas alocadas e planilha salva com sucesso!")
    except Exception as e:
        print(f"Erro ao salvar a planilha: {e}")

if __name__ == "__main__":
    import sys
    if len(sys.argv) < 2:
        print("Uso: python alocarsala.py <caminho_para_planilha>")
    else:
        planilha_path = sys.argv[1]
        alocar_salas(planilha_path)
