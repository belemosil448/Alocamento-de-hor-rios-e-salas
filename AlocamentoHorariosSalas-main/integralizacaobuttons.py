import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
import subprocess
import os

# Variáveis globais
workbook = None
current_sheet = None
planilha_path = None

def escolher_planilha(selected_option, tree):
    global planilha_path
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if file_path:
        planilha_path = file_path
        carregar_planilha(selected_option, tree, file_path)

def carregar_planilha(selected_option, tree, file_path):
    global workbook, current_sheet
    try:
        workbook = load_workbook(file_path)
        print(f"Arquivo {file_path} carregado com sucesso.")
        current_sheet = None  # Limpa a planilha atual ao carregar um novo arquivo
        abrir_planilha_semestre(selected_option, tree)
    except FileNotFoundError:
        print(f"Arquivo {file_path} não encontrado.")
    except Exception as e:
        print(f"Erro ao carregar a planilha: {e}")
        messagebox.showerror("Erro", f"Erro ao carregar a planilha: {e}")

def abrir_planilha_semestre(selected_option, tree):
    global current_sheet
    if workbook is None:
        print("Nenhum arquivo carregado.")
        return
    
    semestre = selected_option.get().replace('° Semestre', '')
    if semestre == "Indefinido":
        sheet_name = "Indefinido"  # Corrige para "Indefinido" como mencionado
    else:
        sheet_name = f"Semestre {semestre}"
    
    if sheet_name in workbook.sheetnames:
        current_sheet = workbook[sheet_name]
        exibir_planilha(tree, current_sheet)
    else:
        print(f"Aba {sheet_name} não encontrada no arquivo.")
        messagebox.showwarning("Aviso", f"Aba {sheet_name} não encontrada no arquivo.")
        # Limpa a visualização da árvore se a aba não for encontrada
        exibir_planilha(tree, None)

def exibir_planilha(tree, sheet):
    if sheet is None:
        # Limpa a visualização da árvore se a planilha não estiver carregada
        for item in tree.get_children():
            tree.delete(item)
        return
    
    for item in tree.get_children():
        tree.delete(item)
    
    headers = [cell.value for cell in sheet[1] if cell.value is not None]
    tree["columns"] = headers
    for col in tree["columns"]:
        tree.heading(col, text=col)
        tree.column(col, width=100)
    
    for row in sheet.iter_rows(min_row=2, values_only=True):
        tree.insert("", tk.END, values=row)

def salvar_planilha():
    global workbook, planilha_path
    if workbook is None:
        print("Nenhuma planilha carregada para salvar.")
        messagebox.showwarning("Aviso", "Nenhuma planilha carregada para salvar.")
        return
    
    if planilha_path:
        try:
            workbook.save(planilha_path)
            print(f"Planilha salva com sucesso em: {planilha_path}")
            messagebox.showinfo("Sucesso", f"Planilha salva com sucesso em: {planilha_path}")
        except Exception as e:
            print(f"Erro ao salvar a planilha: {e}")
            messagebox.showerror("Erro", f"Erro ao salvar a planilha: {e}")
    else:
        print("Caminho da planilha não especificado.")
        messagebox.showwarning("Aviso", "Caminho da planilha não especificado.")

def criar_horarios(selected_option, tree):
    global planilha_path
    if planilha_path is None:
        print("Nenhuma planilha carregada para preencher horários.")
        return
    script_path = os.path.join(os.path.dirname(__file__), 'criarhorarios.py')
    subprocess.run(["python", script_path, planilha_path])
    # Recarregar a planilha após criar horários
    carregar_planilha(selected_option, tree, planilha_path)

def alocar_salas(selected_option, tree):
    global planilha_path
    if planilha_path is None:
        print("Nenhuma planilha carregada para alocar salas.")
        return
    script_path = os.path.join(os.path.dirname(__file__), 'alocarsala.py')
    subprocess.run(["python", script_path, planilha_path])
    # Recarregar a planilha após alocar salas
    carregar_planilha(selected_option, tree, planilha_path)
